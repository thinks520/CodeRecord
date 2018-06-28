#!/usr/bin/python
# -*- coding: UTF-8 -*-

from xml.dom.minidom import parse
import xml.dom.minidom
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook
import csv
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')

wb = Workbook()
# 创建一个sheet
ws = wb.create_sheet()
# 存csv文件
data = []


if __name__ == '__main__':

    # 使用minidom解析器打开 XML 文档
    DOMTree = xml.dom.minidom.parse("12-31-16-cfmd02.xml")
    collection = DOMTree.documentElement
    # 在集合中获取所有file
    files = collection.getElementsByTagName("file")
    # 打印每个MD5值信息的详细信息
    print "*****object infomation*****"

    for file  in files:
        if file.hasAttribute("id"):
            print "id: %s" % file.getAttribute("id")
        md5 = file.getElementsByTagName('md5')[0]
        print "md5: %s" % md5.childNodes[0].data
        sha1 = file.getElementsByTagName('sha1')[0]
        print "sha1: %s" % sha1.childNodes[0].data
        sha256 = file.getElementsByTagName('sha256')[0]
        print "sha256: %s" % sha256.childNodes[0].data
        size = file.getElementsByTagName('size')[0]
        print "size: %s" % size.childNodes[0].data
        filename = file.getElementsByTagName('filename')[0]
        print "filename: %s" % filename.childNodes[0].data
        # 写入xlsx
        ws.append((file.getAttribute("id"),
                         md5.childNodes[0].data,
                         sha1.childNodes[0].data,
                         sha256.childNodes[0].data,
                         size.childNodes[0].data,
                         filename.childNodes[0].data)) #写数据到数据集合中
        # 写入csv
        data.append((file.getAttribute("id"),
                         md5.childNodes[0].data,
                         sha1.childNodes[0].data,
                         sha256.childNodes[0].data,
                         size.childNodes[0].data,
                         filename.childNodes[0].data))


    csvfile = open('test.csv', 'wb')  # 打开方式还可以使用file对象
    writer = csv.writer(csvfile)
    writer.writerow(['id', 'md5', 'sha1', 'sha256', 'size','filename'])
    writer.writerows(data)
    csvfile.close()

    # 写文件到Excel
    wb.save('data_test.xlsx')
